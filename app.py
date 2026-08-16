from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, jsonify
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
import csv
import io
import os
import cloudinary
import cloudinary.uploader
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl import load_workbook
from flask import send_file

load_dotenv()

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET")
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev_key")

DATABASE_URL = os.environ.get("DATABASE_URL")

MAX_INTENTOS = 3
TIEMPO_BLOQUEO_MIN = 15

# ---------------- FOTOS ----------------

UPLOAD_FOLDER = os.path.join("static", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def upload_to_cloudinary(file):
    try:
        result = cloudinary.uploader.upload(
            file,
            folder="reportes_medidores"
        )
        return result.get("secure_url")
    except Exception as e:
        print("Error Cloudinary:", e)
        return None

# ---------------- DB ----------------

def get_db():
    return psycopg2.connect(
        DATABASE_URL,
        sslmode="require",
        cursor_factory=psycopg2.extras.RealDictCursor
    )

# ---------------- HELPERS ----------------

def get_ip():
    return request.headers.get('X-Forwarded-For', request.remote_addr)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        if session.get("rol") != "admin":
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def usuario_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):

        if "usuario" not in session:
            return redirect(url_for("login"))

        if session.get("rol") not in ["usuario", "admin"]:
            abort(403)

        return f(*args, **kwargs)

    return decorated_function

# ---------------- INDEX ----------------

@app.route("/")
def index():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") == "admin":
        return redirect(url_for("dashboard"))

    return render_template("mapa.html", usuario=session["usuario"])

# ---------------- LOGIN ----------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        password = request.form["password"]
        ip = get_ip()

        conn = get_db()
        cur = conn.cursor()

        try:
            cur.execute("SELECT * FROM intentos_login WHERE ip = %s", (ip,))
            intento = cur.fetchone()

            if intento and intento["bloqueado_hasta"]:
                if datetime.now() < intento["bloqueado_hasta"]:
                    flash("IP bloqueada")
                    return redirect(url_for("login"))

            cur.execute("SELECT * FROM usuarios_sistema WHERE usuario = %s", (usuario,))
            user = cur.fetchone()

            if not user:
                registrar_intento(ip, conn)
                flash("Usuario o contraseña incorrectos")
                return redirect(url_for("login"))

            if user["estado"] == "bloqueado":
                flash("Usuario bloqueado")
                return redirect(url_for("login"))

            if not check_password_hash(user["contrasena"], password):
                registrar_intento(ip, conn)
                flash("Usuario o contraseña incorrectos")
                return redirect(url_for("login"))

            limpiar_intentos(ip, conn)

            session["usuario"] = user["usuario"]
            session["rol"] = user["rol"]

            return redirect(url_for("index"))

        finally:
            conn.close()

    return render_template("login.html")

@app.route("/dashboard")
@admin_required
def dashboard():
    return render_template("dashboard.html", usuario=session["usuario"], rol=session["rol"])

# ---------------- LOGOUT ----------------

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ---------------- VER REPORTES ----------------

@app.route("/ver_reportes")
def ver_reportes():
    if "usuario" not in session:
        return redirect(url_for("login"))
    return render_template("ver_reportes.html")

# ---------------- REVISIONES ----------------

@app.route("/admin/revisiones")
@admin_required
def revisiones():

    conn = get_db()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT usuario,nombre
            FROM usuarios_sistema
            WHERE estado='activo'
            ORDER BY nombre
        """)

        usuarios = cur.fetchall()

    finally:
        conn.close()

    return render_template(
    "revisiones.html",
    usuarios=usuarios,
    encontrados=[],
    no_encontrados=[],
    total=0
)


# ---------------- CLIENTES ----------------

@app.route("/api/clientes")
def api_clientes():
    serial = request.args.get("serial")

    conn = get_db()
    cur = conn.cursor()

    try:
        if serial:
            cur.execute("""
                SELECT serialnumber, latitude, longitude, mru
                FROM suscriptores_noviembre_2025
                WHERE serialnumber = %s
            """, (serial,))
        else:
            cur.execute("""
                SELECT serialnumber, latitude, longitude, mru
                FROM suscriptores_noviembre_2025
                LIMIT 200
            """)

        return jsonify(cur.fetchall())

    finally:
        conn.close()

# ---------------- RUTA MRU ----------------

@app.route("/api/mru_ruta")
def api_mru_ruta():

    if "usuario" not in session:
        return jsonify({"error": "No autorizado"}), 401

    mru = request.args.get("mru")

    if not mru:
        return jsonify([])

    conn = get_db()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                serialnumber,
                latitude,
                longitude,
                mru,
                mr_time
            FROM suscriptores_noviembre_2025
            WHERE mru = %s
              AND latitude IS NOT NULL
              AND longitude IS NOT NULL
            ORDER BY mr_time
        """, (mru,))

        datos = cur.fetchall()

        for fila in datos:
            if fila["mr_time"]:
                fila["mr_time"] = fila["mr_time"].strftime("%H:%M:%S")

        return jsonify(datos)

    finally:
        conn.close()
# ---------------- REVISIONES USUARIO ----------------

@app.route("/api/mis_revisiones")
@usuario_required
def api_mis_revisiones():

    conn = get_db()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT COUNT(*) AS cantidad
            FROM ordenes_trabajo
            WHERE usuario=%s
            AND tipo='Revision'
            AND estado='Pendiente'
        """,
        (
            session["usuario"],
        ))

        resultado = cur.fetchone()

        return {
            "cantidad": resultado["cantidad"]
        }

    except Exception as e:

        print("ERROR MIS REVISIONES:", e)

        return {
            "cantidad": 0
        }

    finally:

        conn.close()

# ---------------- RUTA REVISIONES ASIGNADAS ----------------

@app.route("/api/ruta_revisiones")
@usuario_required
def api_ruta_revisiones():

    conn = get_db()
    cur = conn.cursor()

    try:

        cur.execute("""
           SELECT
                id,
                instalacion,
                serialnumber,
                comentario,
                latitud,
                longitud,
                estado
                FROM ordenes_trabajo
            WHERE usuario=%s
            AND tipo='Revision'
            AND estado='Pendiente'
            AND latitud IS NOT NULL
            AND longitud IS NOT NULL
            ORDER BY id
        """,
        (
            session["usuario"],
        ))

        datos = cur.fetchall()

        return jsonify(datos)

    except Exception as e:

        print("ERROR RUTA REVISIONES:", e)

        return jsonify([])

    finally:

        conn.close()

# ---------------------------------------------------------
# GUARDAR EVIDENCIA DE ORDEN DE TRABAJO
# ---------------------------------------------------------

@app.route("/api/ordenes/<int:orden_id>/evidencia", methods=["POST"])
@usuario_required
def guardar_evidencia_orden(orden_id):

    conn = get_db()
    cur = conn.cursor()

    try:

        # ---------------------------------------------
        # VERIFICAR QUE LA ORDEN PERTENECE AL USUARIO
        # ---------------------------------------------

        cur.execute("""
            SELECT
                id,
                estado,
                usuario
            FROM ordenes_trabajo
            WHERE id = %s
              AND usuario = %s
              AND tipo = 'Revision'
        """, (
            orden_id,
            session["usuario"]
        ))

        orden = cur.fetchone()

        if not orden:
            return jsonify({
                "error": "Orden no encontrada"
            }), 404


        # ---------------------------------------------
        # VERIFICAR QUE SIGUE PENDIENTE
        # ---------------------------------------------

        if orden["estado"] != "Pendiente":

            return jsonify({
                "error": "La orden ya fue procesada"
            }), 400


        # ---------------------------------------------
        # DATOS RECIBIDOS
        # ---------------------------------------------

        encontrado = request.form.get("encontrado")

        comentario = request.form.get(
            "comentario",
            ""
        ).strip()

        latitud = request.form.get("latitud")
        longitud = request.form.get("longitud")
         
               # ---------------------------------------------
        # RECIBIR FOTO
        # ---------------------------------------------

        foto_url = None

        if "foto" in request.files:

            archivo = request.files["foto"]

            if archivo and archivo.filename != "":

                if not allowed_file(archivo.filename):

                    return jsonify({
                        "error": "Formato de imagen no permitido"
                    }), 400

                foto_url = upload_to_cloudinary(archivo)

                if not foto_url:

                    return jsonify({
                        "error": "No se pudo subir la fotografía"
                    }), 500
         
        # ---------------------------------------------
        # VALIDAR RESULTADO
        # ---------------------------------------------

        if encontrado == "true":

            encontrado_bool = True

        elif encontrado == "false":

            encontrado_bool = False

        else:

            return jsonify({
                "error": "Debe indicar si el medidor fue encontrado"
            }), 400

        # ---------------------------------------------
        # GUARDAR EVIDENCIA
        # ---------------------------------------------

        cur.execute("""
            INSERT INTO ordenes_trabajo_evidencias
            (
                orden_id,
                comentario,
                foto,
                usuario
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s
            )
        """, (
            orden_id,
            comentario,
            foto_url,
            session["usuario"]
        ))


        # ---------------------------------------------
        # ACTUALIZAR ORDEN
        # ---------------------------------------------

        cur.execute("""
            UPDATE ordenes_trabajo
            SET
                ubicacion_encontrada = %s,
                latitud = %s,
                longitud = %s,
                fecha_inicio = COALESCE(fecha_inicio, NOW()),
                fecha_finalizacion = NOW(),
                estado = 'Completada'
            WHERE id = %s
        """, (
            encontrado_bool,
            latitud,
            longitud,
            orden_id
        ))


        # ---------------------------------------------
        # CONFIRMAR TRANSACCIÓN
        # ---------------------------------------------

        conn.commit()


        print("================================")
        print("EVIDENCIA GUARDADA")
        print("Orden:", orden_id)
        print("Usuario:", session["usuario"])
        print("Encontrado:", encontrado_bool)
        print("Comentario:", comentario)
        print("Latitud:", latitud)
        print("Longitud:", longitud)
        print("Foto:", foto_url)
        print("Estado: Completada")
        print("================================")


        return jsonify({
            "status": "ok",
            "mensaje": "Evidencia guardada correctamente",
            "orden_id": orden_id
        })


    except Exception as e:

        conn.rollback()

        print("ERROR EVIDENCIA:", e)

        return jsonify({
            "error": "Error al guardar evidencia"
        }), 500


    finally:

        conn.close()

# ---------------- REPORTES ----------------

@app.route("/api/reportes", methods=["GET"])
def obtener_reportes():
    if "usuario" not in session:
        return jsonify({"error": "No autorizado"}), 401

    fecha_inicio = request.args.get("inicio")
    fecha_fin = request.args.get("fin")

    conn = get_db()
    cur = conn.cursor()

    try:
        query = """
            SELECT aparato, mru, referencia, comentario, latitud, longitud, usuario, fecha, foto
            FROM reportes_medidores
            WHERE 1=1
        """

        params = []

        if fecha_inicio:
            query += " AND fecha::date >= %s"
            params.append(fecha_inicio)

        if fecha_fin:
            query += " AND fecha::date <= %s"
            params.append(fecha_fin)

        query += " ORDER BY fecha DESC LIMIT 200"

        cur.execute(query, tuple(params))
        return jsonify(cur.fetchall())

    finally:
        conn.close()

# 🔥 AQUÍ ESTÁ LA CORRECCIÓN REAL
@app.route("/api/reportes", methods=["POST"])
def api_reportes():
    if "usuario" not in session:
        return jsonify({"error": "No autorizado"}), 401

    conn = get_db()
    cur = conn.cursor()

    try:
        # Detecta tipo de envío
        if request.content_type and "application/json" in request.content_type:
            data = request.get_json()

            aparato = data.get("aparato")
            mru = data.get("mru")
            referencia = data.get("referencia")
            comentario = data.get("comentario")
            latitud = data.get("latitud")
            longitud = data.get("longitud")
            fecha = data.get("fecha")

            foto_ruta = None

        else:
            aparato = request.form.get("aparato")
            mru = request.form.get("mru")
            referencia = request.form.get("referencia")
            comentario = request.form.get("comentario")
            latitud = request.form.get("latitud")
            longitud = request.form.get("longitud")
            fecha = request.form.get("fecha")

            foto_ruta = None

            if "foto" in request.files:
                archivo = request.files["foto"]

                if archivo and archivo.filename != "" and allowed_file(archivo.filename):

                    # 1️⃣ intentar Cloudinary
                    url_cloud = upload_to_cloudinary(archivo)

                    if url_cloud:
                        foto_ruta = url_cloud  # ☁️ nube
                    else:
                        # 2️⃣ fallback local
                        nombre = secure_filename(archivo.filename)
                        nombre_final = f"{int(datetime.now().timestamp())}_{nombre}"

                        ruta_relativa = os.path.join("uploads", nombre_final)
                        ruta_fisica = os.path.join("static", ruta_relativa)

                        archivo.save(ruta_fisica)

                        foto_ruta = ruta_relativa.replace("\\", "/")

        cur.execute("""
            INSERT INTO reportes_medidores
            (aparato, mru, referencia, comentario, latitud, longitud, usuario, fecha, foto)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            aparato,
            mru,
            referencia,
            comentario,
            latitud,
            longitud,
            session["usuario"],
            fecha,
            foto_ruta
        ))

        conn.commit()

        return jsonify({"status": "ok"})

    except Exception as e:
        print("ERROR:", e)
        return jsonify({"error": "Error al guardar"}), 500

    finally:
        conn.close()
# ---------------- EXPORT ----------------
@app.route("/admin/exportar_reportes")
@admin_required
def exportar_reportes():
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT aparato, mru, referencia, comentario,
                   latitud, longitud, usuario, fecha, foto
            FROM reportes_medidores
            ORDER BY fecha DESC
        """)

        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reportes"

        ws.append([
            "APARATO","MRU","REFERENCIA","COMENTARIO",
            "LAT","LON","USUARIO","FECHA","FOTO"
        ])

        base_url = request.host_url.rstrip("/")

        for r in rows:
            ws.append([
                r["aparato"],
                r["mru"],
                r["referencia"],
                r["comentario"],
                r["latitud"],
                r["longitud"],
                r["usuario"],
                r["fecha"],
                "ver foto" if r["foto"] else ""
            ])

            if r["foto"]:
                fila = ws.max_row
                celda = ws[f"I{fila}"]

                foto = r["foto"]

                if foto:
                    if foto.startswith("http"):
                        url = foto  # Cloudinary
                    else:
                        url = f'{base_url}/static/{foto}'  # local
                else:
                    url = None

                if url:
                    celda.value = "ver foto"
                    celda.hyperlink = url
                    celda.style = "Hyperlink"

        file_path = "reportes.xlsx"
        wb.save(file_path)

        return send_file(file_path, as_attachment=True)

    finally:
        conn.close()
# ---------------- INTENTOS ----------------

def registrar_intento(ip, conn):
    cur = conn.cursor()

    cur.execute("SELECT * FROM intentos_login WHERE ip = %s", (ip,))
    data = cur.fetchone()

    if data:
        intentos = data["intentos"] + 1
        if intentos >= MAX_INTENTOS:
            bloqueo = datetime.now() + timedelta(minutes=TIEMPO_BLOQUEO_MIN)
            cur.execute("""
                UPDATE intentos_login 
                SET intentos=%s, bloqueado_hasta=%s 
                WHERE ip=%s
            """, (intentos, bloqueo, ip))
        else:
            cur.execute("UPDATE intentos_login SET intentos=%s WHERE ip=%s", (intentos, ip))
    else:
        cur.execute("INSERT INTO intentos_login (ip, intentos) VALUES (%s, 1)", (ip,))

    conn.commit()

def limpiar_intentos(ip, conn):
    cur = conn.cursor()
    cur.execute("DELETE FROM intentos_login WHERE ip = %s", (ip,))
    conn.commit()
#---------------------agregado ya que la ia lo elimino -------------
@app.route("/admin/cargar_suscriptores", methods=["GET", "POST"])
@admin_required
def cargar_suscriptores():
    if request.method == "POST":
        archivo = request.files.get("archivo")

        if not archivo:
            flash("No se seleccionó archivo")
            return redirect(url_for("cargar_suscriptores"))

        import tempfile

        contenido = archivo.read().decode("utf-8").lstrip('\ufeff')

        # 🔥 limpiar líneas vacías
        lineas = [l for l in contenido.splitlines() if l.strip() != ""]
        contenido = "\n".join(lineas)  

        # 🔹 guardar temporalmente el CSV
        with tempfile.NamedTemporaryFile(mode='w+', delete=False, suffix=".csv") as temp:
            temp.write(contenido)
            temp_path = temp.name

        conn = get_db()
        cur = conn.cursor()

        try:
            # 🔒 iniciar transacción segura
            conn.autocommit = False

            # 1️⃣ crear tabla temporal
            cur.execute("""
                CREATE TEMP TABLE temp_suscriptores (
    serialnumber TEXT,
    latitude DOUBLE PRECISION,
    longitude DOUBLE PRECISION,
    mru TEXT,
    mr_time TIME
) ON COMMIT DROP;
            """)

            # 2️⃣ cargar datos masivamente
            with open(temp_path, 'r') as f:
                cur.copy_expert("""
                    COPY temp_suscriptores(
    serialnumber,
    latitude,
    longitude,
    mru,
    mr_time
)
FROM STDIN WITH CSV HEADER
                """, f)

            # 3️⃣ limpiar datos inválidos
            cur.execute("""
                DELETE FROM temp_suscriptores
                WHERE latitude = 0 OR longitude = 0 OR serialnumber IS NULL;
            """)

            # 4️⃣ reemplazo seguro
            cur.execute("DELETE FROM suscriptores_noviembre_2025")

            cur.execute("""
                INSERT INTO suscriptores_noviembre_2025 (
    serialnumber,
    latitude,
    longitude,
    mru,
    mr_time
)
SELECT
    serialnumber,
    latitude,
    longitude,
    mru,
    mr_time
FROM temp_suscriptores;
            """)

            # ✅ confirmar todo
            conn.commit()

            flash("Carga masiva completada 🚀")

        except Exception as e:
            conn.rollback()
            print("ERROR:", e)
            flash("Error en la carga")

        finally:
            conn.close()

        return redirect(url_for("cargar_suscriptores"))

    return render_template("cargar_suscriptores.html")
#-----------------------------otro eliminado de la IA -----------------------------------
@app.route("/admin/usuarios")
@admin_required
def admin_usuarios():
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT usuario, nombre, correo, estado, rol
            FROM usuarios_sistema
            ORDER BY usuario
        """)
        usuarios = cur.fetchall()
        return render_template("admin_usuarios.html", usuarios=usuarios)
    finally:
        conn.close()


#------------------- resultados de auditoria --------------

@app.route("/admin/resultados_auditorias")
@admin_required
def resultados_auditorias():

    conn = get_db()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                ot.id,
                ot.instalacion,
                ot.serialnumber,
                ot.usuario,
                ot.estado,
                ot.ubicacion_encontrada,
                ot.latitud,
                ot.longitud,
                ot.fecha_inicio,
                ot.fecha_finalizacion,
                ote.comentario,
                ote.foto,
                ote.fecha AS fecha_evidencia
            FROM ordenes_trabajo ot
            LEFT JOIN LATERAL (
                SELECT
                    comentario,
                    foto,
                    fecha
                FROM ordenes_trabajo_evidencias
                WHERE orden_id = ot.id
                ORDER BY fecha DESC
                LIMIT 1
            ) ote ON TRUE
            WHERE ot.tipo = 'Revision'
            ORDER BY ot.fecha_asignacion DESC, ot.id DESC
        """)

        auditorias = cur.fetchall()


        # =============================================
        # RESUMEN DE AUDITORÍAS POR TÉCNICO
        # =============================================

        cur.execute("""
            SELECT
                usuario,

                COUNT(*) AS asignadas,

                COUNT(*) FILTER (
                    WHERE estado = 'Completada'
                ) AS completadas,

                COUNT(*) FILTER (
                    WHERE estado = 'Pendiente'
                ) AS pendientes,

                COUNT(*) FILTER (
                    WHERE estado = 'Completada'
                    AND ubicacion_encontrada = TRUE
                    ) AS encontrados,

                COUNT(*) FILTER (
                    WHERE estado = 'Completada'
                    AND ubicacion_encontrada = FALSE
                    ) AS no_encontrados
            FROM ordenes_trabajo

            WHERE tipo = 'Revision'

            GROUP BY usuario

            ORDER BY usuario
        """)

        resumen_tecnicos = cur.fetchall()


        return render_template(
            "resultados_auditorias.html",
            auditorias=auditorias,
            resumen_tecnicos=resumen_tecnicos
        )
    
    finally:

        conn.close()
#---------- --------------------------------------------------------

@app.route("/admin/api/ruta_auditorias/<usuario>")
@admin_required
def ruta_auditorias(usuario):

    conn = get_db()
    cur = conn.cursor()

    try:

        cur.execute("""
            SELECT
                ot.id,
                ot.instalacion,
                ot.serialnumber,
                ot.usuario,
                ot.estado,
                ot.ubicacion_encontrada,
                ot.latitud,
                ot.longitud,
                ot.fecha_inicio,
                ot.fecha_finalizacion,
                ote.comentario,
                ote.foto
            FROM ordenes_trabajo ot

            LEFT JOIN LATERAL (
                SELECT
                    comentario,
                    foto
                FROM ordenes_trabajo_evidencias
                WHERE orden_id = ot.id
                ORDER BY fecha DESC
                LIMIT 1
            ) ote ON TRUE

            WHERE ot.tipo = 'Revision'
              AND ot.usuario = %s
              AND ot.latitud IS NOT NULL
              AND ot.longitud IS NOT NULL

            ORDER BY
                COALESCE(
                    ot.fecha_inicio,
                    ot.fecha_finalizacion,
                    ot.fecha_asignacion
                )
        """, (usuario,))

        revisiones = cur.fetchall()

        return jsonify(revisiones)

    finally:

        conn.close()

#------------------ no eliminar del codigo  ----------------

@app.route("/admin/usuario/nuevo", methods=["GET", "POST"])
@admin_required
def nuevo_usuario():
    if request.method == "POST":
        usuario = request.form["usuario"]
        nombre = request.form["nombre"]
        password = request.form["password"]
        correo = request.form["correo"]
        rol = request.form.get("rol", "usuario")

        hash_pw = generate_password_hash(password)

        conn = get_db()
        cur = conn.cursor()

        try:
            cur.execute("SELECT usuario FROM usuarios_sistema WHERE usuario = %s", (usuario,))
            if cur.fetchone():
                flash("Usuario ya existe")
                return redirect(url_for("nuevo_usuario"))

            cur.execute("""
                INSERT INTO usuarios_sistema (usuario, nombre, contrasena, correo, estado, rol)
                VALUES (%s, %s, %s, %s, 'activo', %s)
            """, (usuario, nombre, hash_pw, correo, rol))

            conn.commit()
            flash("Usuario creado")

        finally:
            conn.close()

        return redirect(url_for("admin_usuarios"))

    return render_template("nuevo_usuario.html")

#-------------------------------------
@app.route("/admin/usuario/<usuario>/bloquear")
@admin_required
def bloquear_usuario(usuario):
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE usuarios_sistema
            SET estado = 'bloqueado'
            WHERE usuario = %s
        """, (usuario,))

        conn.commit()
        flash(f"Usuario {usuario} bloqueado.")

    finally:
        conn.close()

    return redirect(url_for("admin_usuarios"))

#-----------------------------------------
@app.route("/admin/usuario/<usuario>/desbloquear")
@admin_required
def desbloquear_usuario(usuario):
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE usuarios_sistema
            SET estado = 'activo'
            WHERE usuario = %s
        """, (usuario,))

        conn.commit()
        flash(f"Usuario {usuario} desbloqueado.")

    finally:
        conn.close()

    return redirect(url_for("admin_usuarios"))

#-------------------------------------------
@app.route('/cambiar_password/<usuario>', methods=['GET', 'POST'])
@admin_required
def cambiar_password(usuario):

    if request.method == 'POST':
        nueva_password = request.form.get('password')

        hash_pw = generate_password_hash(nueva_password)

        conn = get_db()
        cur = conn.cursor()

        try:
            cur.execute("""
                UPDATE usuarios_sistema
                SET contrasena = %s
                WHERE usuario = %s
            """, (hash_pw, usuario))

            conn.commit()
            flash('Contraseña actualizada correctamente')

        finally:
            conn.close()

        return redirect(url_for('admin_usuarios'))

    return render_template('cambiar_password.html', usuario=usuario)

# ---------------- CARGAR REVISIONES ----------------

@app.route("/admin/cargar_revisiones", methods=["POST"])
@admin_required
def cargar_revisiones():

    archivo = request.files.get("archivo")
    usuario = request.form.get("usuario")
    id_lote = f"REV-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    if not archivo:
        flash("Debe seleccionar un archivo Excel.")
        return redirect(url_for("revisiones"))

    try:

        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        encabezados = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        columnas_requeridas = [
            "instalacion",
            "serialnumber",
            "fecha",
            "comentario"
        ]

        for col in columnas_requeridas:

            if col not in encabezados:

                flash(f"No existe la columna: {col}")
                return redirect(url_for("revisiones"))

        idx_inst = encabezados.index("instalacion")
        idx_serial = encabezados.index("serialnumber")
        idx_fecha = encabezados.index("fecha")
        idx_com = encabezados.index("comentario")

        conn = get_db()
        cur = conn.cursor()

        encontrados = []
        no_encontrados = []

        for fila in ws.iter_rows(min_row=2, values_only=True):

            instalacion = fila[idx_inst]
            serial = fila[idx_serial]
            fecha = fila[idx_fecha]
            comentario = fila[idx_com]

            if serial is None:
                continue

            cur.execute("""
                SELECT
                    latitude,
                    longitude
                FROM suscriptores_noviembre_2025
                WHERE serialnumber=%s
            """,(str(serial),))

            dato = cur.fetchone()

            if dato:

                encontrados.append({

                    "instalacion": instalacion,
                    "serialnumber": serial,
                    "fecha": fecha,
                    "comentario": comentario,
                    "latitud": dato["latitude"],
                    "longitud": dato["longitude"]

                })


                cur.execute("""
                    INSERT INTO revisiones_temporal
                    (
                        id_lote,
                        usuario,
                        instalacion,
                        serialnumber,
                        fecha,
                        comentario,
                        latitud,
                        longitud,
                        encontrado
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    id_lote,
                    usuario,
                    instalacion,
                    str(serial),
                    fecha,
                    comentario,
                    dato["latitude"],
                    dato["longitude"],
                    True
                ))

            else:

                no_encontrados.append({

                    "instalacion": instalacion,
                    "serialnumber": serial,
                    "fecha": fecha,
                    "comentario": comentario

                })
                cur.execute("""
                    INSERT INTO revisiones_temporal
                    (
                        id_lote,
                        usuario,
                        instalacion,
                        serialnumber,
                        fecha,
                        comentario,
                        encontrado
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    id_lote,
                    usuario,
                    instalacion,
                    str(serial),
                    fecha,
                    comentario,
                    False
                ))

        conn.commit()
        conn.close()

        # volver a consultar usuarios activos
        conn2 = get_db()
        cur2 = conn2.cursor()

        cur2.execute("""
            SELECT usuario, nombre
            FROM usuarios_sistema
            WHERE estado='activo'
            ORDER BY nombre
        """)

        usuarios = cur2.fetchall()

        conn2.close()

        return render_template(

            "revisiones.html",

            usuarios=usuarios,

            encontrados=encontrados,

            no_encontrados=no_encontrados,

            total=len(encontrados)+len(no_encontrados),

            usuario=usuario,
            id_lote=id_lote

        )

    except Exception as e:

        print(e)

        flash("Error leyendo el archivo Excel")

        return redirect(url_for("revisiones"))

# ---------------- CONFIRMAR REVISION ----------------
# ---------------- CONFIRMAR REVISION ----------------

@app.route("/admin/confirmar_revision/<id_lote>", methods=["POST"])
@admin_required
def confirmar_revision(id_lote):

    conn = get_db()
    cur = conn.cursor()

    try:

        # Obtener todos los registros del lote
        cur.execute("""
            SELECT
                id,
                instalacion,
                serialnumber,
                fecha,
                usuario,
                comentario,
                latitud,
                longitud,
                encontrado
            FROM revisiones_temporal
            WHERE id_lote=%s
            ORDER BY id
        """, (id_lote,))

        registros = cur.fetchall()

        # Actualizar coordenadas de los no encontrados
        # Actualizar coordenadas de los no encontrados
                # Actualizar coordenadas de los no encontrados

        indice_no_encontrado = 1

        for r in registros:

            if not r["encontrado"]:

                lat = request.form.get(f"latitud_{indice_no_encontrado}")
                lon = request.form.get(f"longitud_{indice_no_encontrado}")

                if lat and lon:

                    cur.execute("""
                        UPDATE revisiones_temporal
                        SET
                            latitud=%s,
                            longitud=%s,
                            encontrado=TRUE
                        WHERE id=%s
                    """,
                    (
                        lat,
                        lon,
                        r["id"]
                    ))

                    r["latitud"] = lat
                    r["longitud"] = lon
                    r["encontrado"] = True

                indice_no_encontrado += 1


        # Crear órdenes

        for r in registros:

            if not r["latitud"] or not r["longitud"]:
                continue

            cur.execute("""
                INSERT INTO ordenes_trabajo
                (
                    instalacion,
                    serialnumber,
                    fecha,
                    usuario,
                    tipo,
                    comentario,
                    latitud,
                    longitud,
                    estado,
                    creado_por,
                    ubicacion_encontrada,
                    fecha_asignacion
                )
                VALUES
                (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()
                )
            """,
            (
                r["instalacion"],
                r["serialnumber"],
                r["fecha"],
                r["usuario"],
                "Revision",
                r["comentario"],
                r["latitud"],
                r["longitud"],
                "Pendiente",
                session["usuario"],
                r["encontrado"]
            ))

                # Limpiar tabla temporal después de crear las órdenes

        cur.execute("""
            DELETE FROM revisiones_temporal
            WHERE id_lote=%s
        """, (id_lote,))


        conn.commit()

        flash("Carga confirmada correctamente")

    except Exception as e:

        conn.rollback()
        print("ERROR CONFIRMAR:", e)
        flash("Error al confirmar carga")

    finally:

        conn.close()

    return redirect(url_for("revisiones"))
# ---------------- RUN ----------------

if __name__ == "__main__":
    app.run(debug=True)